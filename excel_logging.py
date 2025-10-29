import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


# --- Constants ---
DEFAULT_FILE_NAME = "measurements.xlsx"
DEFAULT_SHEET_TITLE = "Data"
HEADER_ROW = ["Timestamp", "Value"]
TIMESTAMP_FORMAT = "%Y-%m-%d %H:%M:%S"


def initiate_excel(filename: str = DEFAULT_FILE_NAME) -> None:
    """Creates the Excel file with a header row if it doesn't exist.
    
    Args:
        filename (str): Path to the Excel file. Defaults to DEFAULT_FILE_NAME.
    """
    if not os.path.exists(filename):
        print(f"Creating new Excel file: {filename}")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = DEFAULT_SHEET_TITLE
        sheet.append(HEADER_ROW)
        try:
            workbook.save(filename)
        except PermissionError:
            print(
                f"ERROR: Could not create '{filename}'. Check permissions.",
                file=sys.stderr,
            )


def write_to_excel(value: str | float, filename: str = DEFAULT_FILE_NAME) -> bool:
    """Appends a timestamp and a value to the specified Excel file.

    Args:
        value (str | float): Value to be saved (will be converted to float).
        filename (str): Path to the Excel file. Defaults to DEFAULT_FILE_NAME.

    Returns:
        bool: True if save was successful, otherwise False.
    """
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        print(
            f"ERROR: File not found '{filename}'. Re-initializing",
            file=sys.stderr,
        )
        initiate_excel(filename)
        return False
    except (PermissionError, InvalidFileException):
        print(
            f"ERROR: Could not open '{filename}'. Check if open or corrupted.",
            file=sys.stderr,
        )
        return False
    
    timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)
    try:
        sheet.append([timestamp, float(value)])
        workbook.save(filename)
        print(f"✓ Saved: {timestamp}, {value}")
        return True
    except PermissionError:
        print(
            f"ERROR: Could not save to Excel. Is '{filename}' open?",
            file=sys.stderr,
        )
        return False
    except Exception as e:
        print(
            f"An error occurred while writing to Excel: {e}",
            file=sys.stderr,
        )
        return False

