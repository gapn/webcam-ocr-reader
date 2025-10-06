import cv2
import numpy
import time
import pytesseract
import re
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

def drawLabel(img, text, org, font=cv2.FONT_HERSHEY_SIMPLEX, scale=0.9, color=(0, 255, 255), thick=2):
    """
    Draws a high-contrast label: a filled black rectangle with the text on top.
    """
    (tw, th), base = cv2.getTextSize(text, font, scale, thick)
    x, y = org
    cv2.rectangle(img, (x - 6, y - th - 6), (x + tw + 6, y + base + 6), (0, 0, 0), -1)
    cv2.putText(img, text, (x, y), font, scale, color, thick, cv2.LINE_AA)

def extractNumber(text):
    """
    Uses regex to find the first valid number in a string.
    """
    if not text:
        return None
    
    pattern = r'\s*[-]?\d+[.,]?\d*'
    match = re.search(pattern, text)
    
    if match:
        return match.group(0).strip().replace(',', '.')
    return None 

def processImage(image, scale, isClaheEnabled, clahe, mode, simpleThreshold):
    """
    Applies the full 'scale-first' image processing pipeline to an image.
    """
    roiGray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
    if roiGray.shape[0] == 0 or roiGray.shape[1] == 0:
        return None
    grayScaled = cv2.resize(roiGray, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    
    denoised = cv2.fastNlMeansDenoising(grayScaled, h=10)
    blurred = cv2.GaussianBlur(denoised, (3, 3), 0)
    kernelSharp = numpy.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])
    sharpened = cv2.filter2D(blurred, -1, kernelSharp)
    
    if isClaheEnabled:
        sharpened = clahe.apply(sharpened)
    
    gradientX = cv2.Sobel(sharpened, cv2.CV_64F, 1, 0, ksize=3)
    gradientY = cv2.Sobel(sharpened, cv2.CV_64F, 0, 1, ksize=3)
    gradient = numpy.sqrt(gradientX**2 + gradientY**2)
    gradient = numpy.uint8(gradient / gradient.max() * 255)
    
    enhancedImage = cv2.addWeighted(sharpened, 0.8, gradient, 0.2, 0)            
    
    if mode == 1:
        _, binaryImage = cv2.threshold(enhancedImage, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    elif mode == 2:
        _, binaryImage = cv2.threshold(enhancedImage, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    elif mode == 3:
        binaryImage = cv2.adaptiveThreshold(
            enhancedImage, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            21,
            10
        )
    elif mode == 4:
        binaryImage = cv2.adaptiveThreshold(
            enhancedImage, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY_INV,
            21,
            10
        )
    else:
        _, binaryImage = cv2.threshold(enhancedImage, simpleThreshold, 255, cv2.THRESH_BINARY)
        
    return binaryImage
    
def performOcr(image, psm, extractNumberFunction):
    """
    Performs multi-configuration OCR on a binary image and returns the cleaned number.
    """
    if image is None:
        return None, None
    
    if image.mean() < 127:
        image = cv2.bitwise_not(image)
        
    baseConfiguration = r'--oem 1 -c tessedit_char_whitelist=0123456789.-, -c load_system_dawg=0 -c load_freq_dawg=0'
    
    configurations = [
        f'--psm {psm} {baseConfiguration}',
        f'--psm 7 {baseConfiguration}', #default
        f'--psm 8 {baseConfiguration}', #single word
        f'--psm 13 {baseConfiguration}', #raw line
    ]

    bestOcrText = ""
    for configuration in configurations:
        try:
            text = pytesseract.image_to_string(image, config=configuration).strip()
            if len(text) > len(bestOcrText):
                bestOcrText = text
        except Exception as e:
            print(f"[Tesseract Error]: {e}")
            continue
    
    cleanedOcrText = extractNumber(bestOcrText)
    return cleanedOcrText, bestOcrText

def handleInput(keyPressed, frame, roiCoordinates, mode, simpleThreshold, scale, psm, isClaheEnabled, isMorphologyEnabled, isSaving, saveInterval):
    """
    Handles all keyboard inputs and returns the updated state.
    """
    shouldQuit = False
    
    if keyPressed == ord('q'):
        shouldQuit = True
    elif keyPressed == ord('s'):
        print("Select an ROI and press SPACE or ENTER")
        selection = cv2.selectROI("Webcam OCR - Live", frame, fromCenter=False, showCrosshair=True)
        if selection[2] > 0 and selection[3] > 0:
            roiCoordinates = selection
    elif keyPressed == ord('1'):
        mode = 1
    elif keyPressed == ord('2'):
        mode = 2
    elif keyPressed == ord('3'):
        mode = 3
    elif keyPressed == ord('4'):
        mode = 4
    elif keyPressed == ord('5'):
        mode = 5
    elif keyPressed == ord('['):
        simpleThreshold = max(0, simpleThreshold - 5)
    elif keyPressed == ord(']'):
        simpleThreshold = min(255, simpleThreshold + 5)
    elif keyPressed == ord('c'):
        isClaheEnabled = not isClaheEnabled
    elif keyPressed == ord('m'):
        isMorphologyEnabled = not isMorphologyEnabled
    elif keyPressed in (ord('+'), ord('=')):
        scale = min(8.0, scale + 0.5)
        print(f"Scale set to: {scale}")
    elif keyPressed in (ord('-'), ord('_')):
        scale = max(1.0, scale - 0.5)
        print(f"Scale set to: {scale}")
    elif keyPressed == ord('p'):
        psms = [7, 8, 13, 6]
        current_index = psms.index(psm) if psm in psms else 0
        psm = psms[(current_index + 1) % len(psms)]
        print(f"PSM set to: {psm}")
    elif keyPressed == ord('w'):
        isSaving = not isSaving
        print(f"Saving toggled {'ON' if isSaving else 'OFF'}")
    elif keyPressed == ord(','):
        saveInterval = max(0.5, saveInterval - 0.5)
        print(f"Save interval set to: {saveInterval}s")
    elif keyPressed == ord('.'):
        saveInterval = min(3600.0, saveInterval + 0.5)
        print(f"Save interval set to: {saveInterval}s")
    return shouldQuit, roiCoordinates, mode, simpleThreshold, scale, psm, isClaheEnabled, isMorphologyEnabled, isSaving, saveInterval

def drawOverlays(frame, roiCoordinates, roiCropped, binaryImage, lastOcrText, mode, psm, simpleThreshold, isClaheEnabled, isMorphologyEnabled, fps, isSaving, saveInterval):
    """
    Draws all the text and rectangles on the main frame.
    """
    frameHeight = frame.shape[0]
    
    if roiCoordinates:
        x, y, w, h = roiCoordinates
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
        drawLabel(frame, f"OCR: {lastOcrText or '(empty)'}", (x, max(30, y - 10)))
        
        if roiCropped is not None:
            cv2.imshow("Webcam OCR - ROI", roiCropped)
        if binaryImage is not None:
            processedDisplay = cv2.cvtColor(binaryImage, cv2.COLOR_GRAY2BGR)
            drawLabel(processedDisplay, f"OCR: {lastOcrText or '(empty)'}", (10, 28))
            cv2.imshow("Webcam OCR - Preprocessed ROI", processedDisplay)
            
    else:
        drawLabel(frame, "Press 'S' to select an ROI", (50, frameHeight - 60), scale=0.5)
    
    saveStatus = f"SAVING ({saveInterval}s)" if isSaving else "IDLE"
    saveColor = (0, 255, 255) if isSaving else (50, 220, 50)
    
    hud = (
        f"Mode:{mode} "
        f"PSM: {psm} "
        f"Thr:{simpleThreshold if mode==5 else '-'} "
        f"CLAHE:{'on' if isClaheEnabled else 'off'} "
        f"Morph:{'on' if isMorphologyEnabled else 'off'} "
        f"FPS:{fps:.1f} | {saveStatus}"
    )
    
    drawLabel(frame, hud, (10, frameHeight - 20), scale=0.5, color=saveColor, thick=2)
    
def initExcel(filename="measurements.xlsx"):
    """
    Creates the Excel file with a header row if it doesn't exist.
    """
    if not os.path.exists(filename):
        print(f"Creating new Excel file: {filename}")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        sheet.append(["Timestamp", "Value"])
        workbook.save(filename)

def writeToExcel(value, filename="measurements.xlsx"):
    """
    Appends a timestamp and a value to the specified Excel file.
    """
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([timestamp, float(value)])
        workbook.save(filename)
        print(f"✓ Saved: {timestamp}, {value}")
        return True
    except PermissionError:
        print(f"⚠️ ERROR: Could not save to Excel. Is '{filename}' open?")
        return False
    except Exception as e:
        print(f"⚠️ An error occurred while writing to Excel: {e}")
        return False

def main() -> None:
    """
    Open default webcam, draw ROI rect on live feed and show full frame and cropped ROI in separate windows.
    Press 'q' to quit.
    Glossary:
      - ROI: Region of Interest. The box you draw on the screen.
      - OCR: Optical Character Recognition. The process of reading text from an image.
      - PSM: Page Segmentation Mode. A Tesseract setting for how it views the image.
      - CLAHE: Contrast Limited Adaptive Histogram Equalization. An algorithm to improve image contrast.
      - FPS: Frames Per Second.
    """
    
    videoCapture = cv2.VideoCapture(0)

    initExcel()
    
    roiCoordinates = None
    
    if not videoCapture.isOpened():
        raise RuntimeError("Could not open webcam (index 0). Try different index: 1, 2, ...")
    
    # Loop paramethers
    mode = 1
    simpleThreshold = 100
    scale = 2.5
    psm = 7
    isClaheEnabled = True
    isMorphologyEnabled = False
    isSaving = False
    saveInterval = 5.0
    
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    
    # OCR and FPS controls
    ocrIntervalSeconds = 0.2
    lastOcrTime = 0.0
    lastSaveTime = 0.0
    lastOcrText = ""
    fps = 0.0
    frameCount = 0
    fpsTimer = time.time()
    
    while True:
        isFrameRead, frame = videoCapture.read()
        now = time.time()
        
        if not isFrameRead:
            break

        roiCropped = None
        binaryImage = None
        
        if roiCoordinates:
            x, y, w, h = roiCoordinates
            roiCropped = frame[y:y+h, x:x+w]
        
            binaryImage = processImage(
                roiCropped, scale, isClaheEnabled, clahe, mode, simpleThreshold
            )
                
            if binaryImage is not None and (now - lastOcrTime >= ocrIntervalSeconds):
                newOcrText, rawOcrText = performOcr(binaryImage, psm, extractNumber)
                
                #print("OCR raw    :", repr(rawOcrText))
                #print("OCR clean  :", repr(newOcrText))

                if newOcrText:
                    lastOcrText = newOcrText
                    
                lastOcrTime = now
            
            if isSaving and lastOcrText and (now - lastSaveTime) >= saveInterval:
                if writeToExcel(lastOcrText):
                    lastSaveTime = now
        
        frameCount += 1
        elapsed = now - fpsTimer
        if elapsed >= 1.0:
            fps = frameCount / elapsed
            frameCount = 0
            fpsTimer = now
            
        drawOverlays(frame, roiCoordinates, roiCropped, binaryImage, lastOcrText, mode, psm, simpleThreshold, isClaheEnabled, isMorphologyEnabled, fps, isSaving, saveInterval)

        cv2.imshow("Webcam OCR - Live", frame)
        
        keyPressed = cv2.waitKey(1) & 0xFF
        shouldQuit, roiCoordinates, mode, simpleThreshold, scale, psm, isClaheEnabled, isMorphologyEnabled, isSaving, saveInterval = handleInput(
            keyPressed, frame, roiCoordinates, mode, simpleThreshold, scale, psm, isClaheEnabled, isMorphologyEnabled, isSaving, saveInterval
        )
        if shouldQuit:
            break
        
    print("Closing application...")
    videoCapture.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    main()